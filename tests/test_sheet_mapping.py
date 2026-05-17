"""Unit tests cho sheet_mapping module."""

from __future__ import annotations

import pytest

from auto_fill.mapper.sheet_mapping import (
    AUTO,
    BHYT_BHXH,
    HEALTH,
    MOTORBIKE,
    SHEET_COL_MAPS,
    STUDENT,
    TRAVEL,
    get_map,
)


class TestSheetColumnMapStructure:
    def test_all_aliases_registered(self) -> None:
        assert set(SHEET_COL_MAPS.keys()) == {
            "travel",
            "auto",
            "motorbike",
            "health",
            "bhyt_bhxh",
            "student",
        }

    def test_get_map_returns_correct_instance(self) -> None:
        assert get_map("travel") is TRAVEL
        assert get_map("auto") is AUTO
        assert get_map("motorbike") is MOTORBIKE
        assert get_map("health") is HEALTH
        assert get_map("bhyt_bhxh") is BHYT_BHXH
        assert get_map("student") is STUDENT

    def test_get_map_unknown_alias_raises(self) -> None:
        with pytest.raises(KeyError):
            get_map("nonexistent")

    def test_all_maps_are_frozen(self) -> None:
        for m in SHEET_COL_MAPS.values():
            with pytest.raises((AttributeError, TypeError)):
                m.alias = "x"  # type: ignore[misc]

    @pytest.mark.parametrize(
        "alias", ["travel", "auto", "motorbike", "health", "bhyt_bhxh", "student"]
    )
    def test_col_indices_are_positive(self, alias: str) -> None:
        m = get_map(alias)
        for _, col_idx in m.col_map:
            assert col_idx >= 1, f"col_idx {col_idx} must be >= 1 in {alias}"

    @pytest.mark.parametrize(
        "alias", ["travel", "auto", "motorbike", "health", "bhyt_bhxh", "student"]
    )
    def test_no_duplicate_col_indices(self, alias: str) -> None:
        m = get_map(alias)
        col_indices = [col for _, col in m.col_map]
        assert len(col_indices) == len(set(col_indices)), f"Duplicate col indices in {alias}"


class TestTravelMapping:
    def test_stt_col_is_2(self) -> None:
        assert TRAVEL.stt_col == 2

    def test_issued_date_col_1(self) -> None:
        assert ("issued_date", 1) in TRAVEL.col_map

    def test_insured_name_col_3(self) -> None:
        assert ("insured_name", 3) in TRAVEL.col_map

    def test_premium_col_14(self) -> None:
        assert ("premium", 14) in TRAVEL.col_map

    def test_trip_days_is_computed(self) -> None:
        assert "trip_days" in TRAVEL.computed_fields

    def test_stt_entry_is_none(self) -> None:
        none_cols = [col for field, col in TRAVEL.col_map if field is None]
        assert 2 in none_cols  # STT


class TestAutoMapping:
    def test_plate_number_col_7(self) -> None:
        assert ("plate_number", 7) in AUTO.col_map

    def test_stt_col_is_2(self) -> None:
        assert AUTO.stt_col == 2


class TestMotorbikeMapping:
    def test_plate_number_col_3(self) -> None:
        assert ("plate_number", 3) in MOTORBIKE.col_map

    def test_effective_from_col_12(self) -> None:
        assert ("effective_from", 12) in MOTORBIKE.col_map

    def test_premium_col_10(self) -> None:
        assert ("premium", 10) in MOTORBIKE.col_map


class TestHealthMapping:
    def test_insured_dob_col_4(self) -> None:
        # NĐBH DOB — must be col 4 (DATABASE §3.2)
        assert ("insured_dob", 4) in HEALTH.col_map

    def test_insured_id_number_col_8(self) -> None:
        assert ("insured_id_number", 8) in HEALTH.col_map

    def test_buyer_name_col_10(self) -> None:
        assert ("buyer_name", 10) in HEALTH.col_map

    def test_buyer_relation_col_11(self) -> None:
        assert ("buyer_relation", 11) in HEALTH.col_map

    def test_buyer_dob_col_12(self) -> None:
        # BMBH DOB — must be col 12, DISTINCT from insured_dob (col 4)
        assert ("buyer_dob", 12) in HEALTH.col_map

    def test_insured_dob_and_buyer_dob_are_different_columns(self) -> None:
        col_map = dict(HEALTH.col_map)
        assert col_map.get("insured_dob") != col_map.get("buyer_dob")
        assert col_map["insured_dob"] == 4
        assert col_map["buyer_dob"] == 12


class TestBhytBhxhMapping:
    def test_insured_dob_col_4(self) -> None:
        assert ("insured_dob", 4) in BHYT_BHXH.col_map

    def test_buyer_dob_col_11(self) -> None:
        # BMBH DOB at col 11, DISTINCT from insured_dob (col 4)
        assert ("buyer_dob", 11) in BHYT_BHXH.col_map

    def test_insured_dob_and_buyer_dob_are_different_columns(self) -> None:
        col_map = dict(BHYT_BHXH.col_map)
        assert col_map["insured_dob"] == 4
        assert col_map["buyer_dob"] == 11

    def test_insured_id_number_col_6(self) -> None:
        assert ("insured_id_number", 6) in BHYT_BHXH.col_map

    def test_buyer_relation_col_13(self) -> None:
        assert ("buyer_relation", 13) in BHYT_BHXH.col_map


class TestStudentMapping:
    def test_insured_dob_col_3(self) -> None:
        assert ("insured_dob", 3) in STUDENT.col_map

    def test_buyer_dob_col_13(self) -> None:
        # BMBH DOB at col 13, DISTINCT from insured_dob (col 3)
        assert ("buyer_dob", 13) in STUDENT.col_map

    def test_insured_dob_and_buyer_dob_are_different_columns(self) -> None:
        col_map = dict(STUDENT.col_map)
        assert col_map["insured_dob"] == 3
        assert col_map["buyer_dob"] == 13

    def test_school_col_9(self) -> None:
        assert ("school", 9) in STUDENT.col_map

    def test_class_name_col_10(self) -> None:
        assert ("class_name", 10) in STUDENT.col_map

    def test_no_stt_col(self) -> None:
        assert STUDENT.stt_col is None


class TestDualDobAcceptance:
    """Acceptance tests cho 8.2b: đảm bảo insured_dob ≠ buyer_dob khi ghi vào master."""

    def _col_for(self, alias: str, field: str) -> int:
        """Trả về col index của field trong sheet alias."""
        m = get_map(alias)
        return dict(m.col_map)[field]

    def test_health_insured_dob_col_4_buyer_dob_col_12(self) -> None:
        assert self._col_for("health", "insured_dob") == 4
        assert self._col_for("health", "buyer_dob") == 12

    def test_bhyt_bhxh_insured_dob_col_4_buyer_dob_col_11(self) -> None:
        assert self._col_for("bhyt_bhxh", "insured_dob") == 4
        assert self._col_for("bhyt_bhxh", "buyer_dob") == 11

    def test_student_insured_dob_col_3_buyer_dob_col_13(self) -> None:
        assert self._col_for("student", "insured_dob") == 3
        assert self._col_for("student", "buyer_dob") == 13

    def test_self_buyer_pattern_same_dob_goes_to_both_columns(self) -> None:
        """Pattern 1: buyer_relation='Bản thân' — insured_dob == buyer_dob value."""
        from datetime import date

        import openpyxl

        wb = openpyxl.Workbook()
        assert wb.active is not None
        ws = wb.active
        ws.title = "Sức khỏe"
        ws.append([""] * 20)

        dob = date(1995, 8, 28)
        record = {
            "insured_name": "Nguyễn Thị Tuyết Linh",
            "insured_dob": dob,
            "buyer_name": "Nguyễn Thị Tuyết Linh",
            "buyer_dob": dob,
            "buyer_relation": "Bản thân",
        }

        row_idx = ws.max_row + 1
        m = get_map("health")
        for canonical, col_idx in m.col_map:
            if canonical and canonical in record:
                ws.cell(row=row_idx, column=col_idx, value=record[canonical])

        assert ws.cell(row=row_idx, column=4).value == dob  # insured_dob
        assert ws.cell(row=row_idx, column=12).value == dob  # buyer_dob
        assert ws.cell(row=row_idx, column=4).value == ws.cell(row=row_idx, column=12).value

    def test_pattern2_different_dob_goes_to_different_columns(self) -> None:
        """Pattern 2: mẹ Thuận mua cho con Phúc — col 4 ≠ col 12."""
        from datetime import date

        import openpyxl

        wb = openpyxl.Workbook()
        assert wb.active is not None
        ws = wb.active
        ws.title = "Sức khỏe"
        ws.append([""] * 20)

        child_dob = date(2018, 12, 12)
        mother_dob = date(1988, 8, 30)
        record = {
            "insured_name": "Đỗ Danh Minh Phúc",
            "insured_dob": child_dob,
            "buyer_name": "Đỗ Thị Thuận",
            "buyer_dob": mother_dob,
            "buyer_relation": "Mẹ",
        }

        row_idx = ws.max_row + 1
        m = get_map("health")
        for canonical, col_idx in m.col_map:
            if canonical and canonical in record:
                ws.cell(row=row_idx, column=col_idx, value=record[canonical])

        assert ws.cell(row=row_idx, column=4).value == child_dob  # insured_dob
        assert ws.cell(row=row_idx, column=12).value == mother_dob  # buyer_dob
        assert ws.cell(row=row_idx, column=4).value != ws.cell(row=row_idx, column=12).value
