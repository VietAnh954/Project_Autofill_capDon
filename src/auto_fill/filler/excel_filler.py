"""Excel filler — append rows vao sheet Du lich trong file tong.

Mapping canonical → column index theo MAPPING.md §6.1.
STT tu dong tang, date format dd/mm/yyyy, money format so.
Pattern 3: rows sharing source_buyer_group_id → only first row writes buyer fields.
"""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from auto_fill.mapper.sheet_mapping import SheetColumnMap

logger = logging.getLogger(__name__)

# MAPPING §6.1 — canonical field → 1-indexed column in sheet Du lich
# None = computed internally (STT, trip_days)
_TRAVEL_COL_MAP: list[tuple[str | None, int]] = [
    ("issued_date", 1),
    (None, 2),  # STT — auto
    ("insured_name", 3),
    ("insured_dob", 4),
    ("insured_id_number", 5),
    ("effective_from", 6),
    ("trip_start", 7),
    ("trip_end", 8),
    (None, 9),  # trip_days — computed
    ("destination", 10),
    ("scope", 11),
    ("plan", 13),
    ("premium", 14),
    ("buyer_name", 15),
]

_DATE_NUMBER_FORMAT = "DD/MM/YYYY"

# Canonical buyer field names — skipped for Pattern 3 continuation rows
_BUYER_FIELDS: frozenset[str] = frozenset(
    {
        "buyer_name",
        "buyer_id_number",
        "buyer_dob",
        "buyer_phone",
        "buyer_email",
        "buyer_address",
        "buyer_gender",
        "buyer_relation",
    }
)


def append_travel_rows(
    records: list[dict[str, Any]],
    master_path: Path,
    sheet_name: str = "Du lịch",
    header_row: int = 1,
) -> int:
    """Append records vao sheet Du lich.

    Args:
        records: Danh sach dict canonical da qua normalizer + validator.
        master_path: Duong dan file tong (.xlsx). FILE PHAI TON TAI.
        sheet_name: Ten sheet trong file tong.
        header_row: Row index (1-based) cua header. Data bat dau o header_row+1.

    Returns:
        So dong da ghi thanh cong.

    Raises:
        FileNotFoundError: Neu master_path khong ton tai.
        KeyError: Neu sheet_name khong ton tai trong workbook.
    """
    if not master_path.exists():
        raise FileNotFoundError(f"File tong khong ton tai: {master_path}")

    wb = openpyxl.load_workbook(master_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet {sheet_name!r} khong ton tai trong {master_path.name}")

    ws: Worksheet = wb[sheet_name]
    next_stt = _next_stt(ws, col_idx=2, header_row=header_row)

    seen_group_ids: set[str] = set()
    written = 0
    for record in records:
        group_id = record.get("source_buyer_group_id")
        skip_buyer = bool(group_id and group_id in seen_group_ids)
        if group_id:
            seen_group_ids.add(str(group_id))
        _write_row(ws, record, next_stt, skip_buyer=skip_buyer)
        next_stt += 1
        written += 1
        logger.info(
            "row_appended",
            extra={
                "sheet": sheet_name,
                "id": record.get("insured_id_number", ""),
                "stt": next_stt - 1,
            },
        )

    wb.save(master_path)
    logger.info("workbook_saved", extra={"path": str(master_path), "rows_written": written})
    return written


def append_rows(
    records: list[dict[str, Any]],
    col_map: SheetColumnMap,
    master_path: Path,
    sheet_name: str,
    header_row: int = 1,
) -> int:
    """Generic append for any sheet with Pattern 3 buyer-empty preservation.

    Records sharing the same source_buyer_group_id: only the first row writes
    buyer fields (cols mapped to buyer_* canonicals). Subsequent rows in the
    group leave those cells empty, matching the original Excel source pattern.

    Args:
        records: Canonical dicts from normalizer/validator.
        col_map: SheetColumnMap for target sheet.
        master_path: Path to master workbook (must exist).
        sheet_name: Sheet name within workbook.
        header_row: 1-based header row index.

    Returns:
        Number of rows written.
    """
    if not master_path.exists():
        raise FileNotFoundError(f"File tong khong ton tai: {master_path}")

    wb = openpyxl.load_workbook(master_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet {sheet_name!r} khong ton tai trong {master_path.name}")

    ws: Worksheet = wb[sheet_name]
    next_stt = (
        _next_stt(ws, col_idx=col_map.stt_col, header_row=header_row) if col_map.stt_col else None
    )

    seen_group_ids: set[str] = set()
    written = 0
    for record in records:
        group_id = record.get("source_buyer_group_id")
        skip_buyer = bool(group_id and group_id in seen_group_ids)
        if group_id:
            seen_group_ids.add(str(group_id))

        _write_row_generic(ws, record, col_map, stt=next_stt, skip_buyer=skip_buyer)
        if next_stt is not None:
            next_stt += 1
        written += 1
        logger.info(
            "row_appended",
            extra={
                "sheet": sheet_name,
                "id": record.get("insured_id_number", ""),
                "skip_buyer": skip_buyer,
            },
        )

    wb.save(master_path)
    logger.info("workbook_saved", extra={"path": str(master_path), "rows_written": written})
    return written


def _next_stt(ws: Worksheet, col_idx: int, header_row: int) -> int:
    """Tinh STT tiep theo = max(cot STT) + 1, min = 1."""
    max_stt = 0
    for row in ws.iter_rows(
        min_row=header_row + 1,
        min_col=col_idx,
        max_col=col_idx,
        values_only=True,
    ):
        val = row[0]
        if isinstance(val, int | float) and val > max_stt:
            max_stt = int(val)
    return max_stt + 1


def _write_row(
    ws: Worksheet, record: dict[str, Any], stt: int, *, skip_buyer: bool = False
) -> None:
    """Ghi 1 record vao dong moi cuoi ws (travel sheet)."""
    next_row = ws.max_row + 1

    # STT
    ws.cell(row=next_row, column=2, value=stt)

    # trip_days — computed
    trip_start: date | None = record.get("trip_start")
    trip_end: date | None = record.get("trip_end")
    if isinstance(trip_start, date) and isinstance(trip_end, date):
        trip_days = (trip_end - trip_start).days
        ws.cell(row=next_row, column=9, value=trip_days)

    # Mapped fields
    for field, col_idx in _TRAVEL_COL_MAP:
        if field is None:
            continue  # handled separately (STT, trip_days)
        if skip_buyer and field in _BUYER_FIELDS:
            continue
        value = record.get(field)
        if value is None:
            continue
        cell = ws.cell(row=next_row, column=col_idx, value=_coerce(value))
        if isinstance(value, date):
            cell.number_format = _DATE_NUMBER_FORMAT
            cell.alignment = Alignment(horizontal="center")


def _write_row_generic(
    ws: Worksheet,
    record: dict[str, Any],
    col_map: SheetColumnMap,
    stt: int | None,
    *,
    skip_buyer: bool = False,
) -> None:
    """Ghi 1 record dung SheetColumnMap bat ky."""
    next_row = ws.max_row + 1

    if stt is not None and col_map.stt_col is not None:
        ws.cell(row=next_row, column=col_map.stt_col, value=stt)

    _maybe_write_trip_days(ws, record, col_map, next_row)

    for field, col_idx in col_map.col_map:
        if field is None:
            continue
        if skip_buyer and field in _BUYER_FIELDS:
            continue
        value = record.get(field)
        if value is None:
            continue
        cell = ws.cell(row=next_row, column=col_idx, value=_coerce(value))
        if isinstance(value, date):
            cell.number_format = _DATE_NUMBER_FORMAT
            cell.alignment = Alignment(horizontal="center")


def _maybe_write_trip_days(
    ws: Worksheet, record: dict[str, Any], col_map: SheetColumnMap, next_row: int
) -> None:
    if "trip_days" not in col_map.computed_fields:
        return
    trip_start = record.get("trip_start")
    trip_end = record.get("trip_end")
    if not (isinstance(trip_start, date) and isinstance(trip_end, date)):
        return
    trip_days = (trip_end - trip_start).days
    for f, col_idx in col_map.col_map:
        if f is None and col_idx == 9:
            ws.cell(row=next_row, column=col_idx, value=trip_days)
            break


def _coerce(value: Any) -> Any:
    """Chuyen gia tri thanh kieu Excel-friendly."""
    if isinstance(value, date):
        # openpyxl nhan datetime.date truc tiep
        return value
    if isinstance(value, float):
        return value
    return str(value) if value is not None else None
