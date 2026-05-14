"""FastAPI application — query, export, web dashboard, pipeline control.

Chay:
    uvicorn auto_fill.api.app:app --reload
    python -m auto_fill serve

Endpoints (API):
    GET  /health                     — kiem tra ket noi DB
    GET  /stats                      — dem ban ghi moi sheet
    GET  /records/{sheet}            — list ban ghi (skip, limit, filter)
    GET  /records/{sheet}/{id}       — lay 1 ban ghi theo id
    GET  /export/{sheet}             — tai Excel snapshot 1 sheet
    GET  /export                     — tai Excel snapshot tat ca sheet

Endpoints (pipeline control):
    POST /pipeline/run               — kich hoat pipeline trong background
    GET  /pipeline/status            — trang thai hien tai
    GET  /pipeline/logs/stream       — SSE stream log realtime

Endpoints (dashboard):
    GET  /                           — redirect sang /dashboard
    GET  /dashboard                  — trang tong quan
    GET  /dashboard/records          — trang duyet ho so
"""

from __future__ import annotations

import asyncio
import io
import threading
import time
from datetime import date
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException, Query
from fastapi.requests import Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.orm import Session

from auto_fill.api.pipeline_state import run_pipeline
from auto_fill.api.pipeline_state import state as pipeline_state
from auto_fill.api.schemas import HealthOut, RecordOut, StatsOut
from auto_fill.db.engine import get_db_url_from_settings, get_engine
from auto_fill.db.models import (
    AutoInsurance,
    BhytBhxhInsurance,
    HealthInsurance,
    MotorbikeInsurance,
    StudentInsurance,
    TravelInsurance,
)

_TEMPLATES_DIR = Path(__file__).parent / "templates"
templates = Jinja2Templates(directory=str(_TEMPLATES_DIR))

# --------------------------------------------------------------------------- #
# Master file cache (Google Drive)                                             #
# --------------------------------------------------------------------------- #

_master_cache: dict[str, Any] = {"xlsx_path": None, "timestamp": 0.0, "error": None}
_cache_lock = threading.Lock()


def _get_master_xlsx(force: bool = False) -> tuple[Path | None, float, str | None]:
    """Return (xlsx_path, timestamp, warning).

    Downloads from Google Drive if cache is stale or force=True.
    Returns stale cache + warning if download fails but old file exists.
    """
    from auto_fill.config.settings import settings

    file_id = getattr(settings, "gdrive_master_file_id", "")
    creds_file = getattr(settings, "gdrive_credentials_file", None)
    if not file_id or not creds_file:
        return None, 0.0, "Google Drive chua duoc cau hinh. Chay: python -m auto_fill drive-setup"

    with _cache_lock:
        ttl = getattr(settings, "gdrive_cache_minutes", 5) * 60
        age = time.time() - _master_cache["timestamp"]
        xlsx_path: Path | None = _master_cache["xlsx_path"]

        if not force and xlsx_path and xlsx_path.exists() and age < ttl:
            return xlsx_path, _master_cache["timestamp"], None

        try:
            from auto_fill.drive.gdrive_client import DriveClient

            token_file = getattr(settings, "gdrive_token_file", None) or (
                Path(str(creds_file)).parent / "gdrive_token.json"
            )
            dest = settings.data_root / "master" / "master_gdrive_cache.xlsx"
            client = DriveClient(str(creds_file), str(token_file))
            client.download_as_xlsx(file_id, dest)

            _master_cache["xlsx_path"] = dest
            _master_cache["timestamp"] = time.time()
            _master_cache["error"] = None
            return dest, _master_cache["timestamp"], None

        except Exception as exc:
            err_msg = str(exc)
            _master_cache["error"] = err_msg
            stale = _master_cache["xlsx_path"]
            if stale and Path(stale).exists():
                return stale, _master_cache["timestamp"], f"Du lieu cu (loi cap nhat: {err_msg})"
            return None, 0.0, err_msg


app = FastAPI(
    title="AutoFill Cap Don API",
    description="Query, export va dieu khien pipeline phieu cap don bao hiem.",
    version="0.6.0",
)

_ALIAS_TO_MODEL: dict[str, type[Any]] = {
    "travel": TravelInsurance,
    "auto": AutoInsurance,
    "motorbike": MotorbikeInsurance,
    "health": HealthInsurance,
    "bhyt_bhxh": BhytBhxhInsurance,
    "student": StudentInsurance,
}

_DATE_FIELD: dict[str, str] = {
    "travel": "issued_date",
    "auto": "issued_date",
    "motorbike": "issued_date",
    "health": "effective_from",
    "bhyt_bhxh": "effective_from",
    "student": "issued_date",
}


def _get_engine() -> Any:
    return get_engine(get_db_url_from_settings())


def _row_to_dict(row: Any) -> dict[str, Any]:
    mapper = type(row).__mapper__
    return {col.key: getattr(row, col.key, None) for col in mapper.columns}


# --------------------------------------------------------------------------- #
# Dashboard routes                                                             #
# --------------------------------------------------------------------------- #


@app.get("/", include_in_schema=False)
def root_redirect() -> RedirectResponse:
    return RedirectResponse(url="/dashboard")


@app.get("/dashboard", response_class=HTMLResponse, include_in_schema=False)
def dashboard(request: Request) -> Any:  # type: ignore[type-arg]
    return templates.TemplateResponse(request, "dashboard.html", {"active_page": "dashboard"})


@app.get("/dashboard/records", response_class=HTMLResponse, include_in_schema=False)
def dashboard_records(request: Request) -> Any:  # type: ignore[type-arg]
    return templates.TemplateResponse(request, "records.html", {"active_page": "records"})


# --------------------------------------------------------------------------- #
# Pipeline control                                                             #
# --------------------------------------------------------------------------- #


class RunRequest(BaseModel):
    dry_run: bool = False


@app.post("/pipeline/run")
def pipeline_run(req: RunRequest) -> dict[str, Any]:
    """Kich hoat pipeline trong background thread."""
    started = run_pipeline(dry_run=req.dry_run)
    if started:
        return {"started": True, "message": "Pipeline đã khởi động."}
    return {"started": False, "message": "Pipeline đang chạy, vui lòng đợi."}


@app.get("/pipeline/status")
def pipeline_status() -> dict[str, Any]:
    """Tra ve trang thai hien tai cua pipeline."""
    return {
        "status": pipeline_state.status,
        "started_at": pipeline_state.started_at.isoformat() if pipeline_state.started_at else None,
        "finished_at": pipeline_state.finished_at.isoformat()
        if pipeline_state.finished_at
        else None,
        "error": pipeline_state.error,
        "log_count": len(pipeline_state.logs),
    }


@app.get("/pipeline/logs/stream")
async def pipeline_logs_stream() -> StreamingResponse:
    """SSE stream: day log moi cua pipeline den trinh duyet."""

    async def generate() -> Any:
        last_sent = 0
        while True:
            logs = list(pipeline_state.logs)
            new_lines = logs[last_sent:]
            for line in new_lines:
                # Escape newlines inside the data payload
                safe = line.replace("\n", " ")
                yield f"data: {safe}\n\n"
            last_sent = len(logs)
            # Heartbeat so the client connection stays alive
            yield "data: __heartbeat__\n\n"
            await asyncio.sleep(1)

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


# --------------------------------------------------------------------------- #
# API routes (existing)                                                        #
# --------------------------------------------------------------------------- #


@app.get("/health", response_model=HealthOut)
def health() -> HealthOut:
    """Kiem tra ket noi DB."""
    db_url = get_db_url_from_settings()
    engine = get_engine(db_url)
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return HealthOut(status="ok", db_url=_mask_url(db_url))
    except Exception as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc


@app.get("/stats", response_model=StatsOut)
def stats() -> StatsOut:
    """Tong so ban ghi moi sheet."""
    engine = _get_engine()
    counts: dict[str, int] = {}
    with Session(engine) as s:
        for alias, model_cls in _ALIAS_TO_MODEL.items():
            counts[alias] = s.query(model_cls).count()
    return StatsOut(
        travel=counts.get("travel", 0),
        auto=counts.get("auto", 0),
        motorbike=counts.get("motorbike", 0),
        health=counts.get("health", 0),
        bhyt_bhxh=counts.get("bhyt_bhxh", 0),
        student=counts.get("student", 0),
        total=sum(counts.values()),
    )


@app.get("/records/{sheet}", response_model=list[RecordOut])
def list_records(
    sheet: str,
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=1000),
    date_from: date | None = None,
    date_to: date | None = None,
    insured_name: str | None = None,
) -> list[RecordOut]:
    """List ban ghi cua 1 sheet voi filter."""
    if sheet not in _ALIAS_TO_MODEL:
        raise HTTPException(status_code=404, detail=f"Sheet '{sheet}' khong ton tai.")
    model_cls = _ALIAS_TO_MODEL[sheet]
    date_col = _DATE_FIELD[sheet]
    engine = _get_engine()

    with Session(engine) as s:
        q = s.query(model_cls)
        if date_from:
            q = q.filter(getattr(model_cls, date_col) >= date_from)
        if date_to:
            q = q.filter(getattr(model_cls, date_col) <= date_to)
        if insured_name:
            name_attr = getattr(model_cls, "insured_name", None)
            if name_attr is not None:
                q = q.filter(name_attr.ilike(f"%{insured_name}%"))
        rows = q.order_by(model_cls.id.desc()).offset(skip).limit(limit).all()
    return [RecordOut(id=row.id, sheet=sheet, data=_row_to_dict(row)) for row in rows]


@app.get("/records/{sheet}/{record_id}", response_model=RecordOut)
def get_record(sheet: str, record_id: int) -> RecordOut:
    """Lay 1 ban ghi theo id."""
    if sheet not in _ALIAS_TO_MODEL:
        raise HTTPException(status_code=404, detail=f"Sheet '{sheet}' khong ton tai.")
    model_cls = _ALIAS_TO_MODEL[sheet]
    engine = _get_engine()

    with Session(engine) as s:
        row = s.get(model_cls, record_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Ban ghi {record_id} khong tim thay.")
    return RecordOut(id=row.id, sheet=sheet, data=_row_to_dict(row))


@app.get("/export/{sheet}")
def export_sheet(sheet: str) -> StreamingResponse:
    """Tai Excel snapshot cua 1 sheet."""
    if sheet not in _ALIAS_TO_MODEL:
        raise HTTPException(status_code=404, detail=f"Sheet '{sheet}' khong ton tai.")
    return _export_response([sheet], f"export_{sheet}.xlsx")


@app.get("/export")
def export_all() -> StreamingResponse:
    """Tai Excel snapshot tat ca 6 sheet."""
    return _export_response(list(_ALIAS_TO_MODEL.keys()), "export_all.xlsx")


# --------------------------------------------------------------------------- #
# Master file routes (Google Drive viewer)                                     #
# --------------------------------------------------------------------------- #


@app.get("/master/status")
def master_status() -> dict[str, Any]:
    """Kiem tra cau hinh Google Drive va trang thai cache."""
    from auto_fill.config.settings import settings

    file_id = getattr(settings, "gdrive_master_file_id", "")
    creds_file = getattr(settings, "gdrive_credentials_file", None)
    configured = bool(file_id) and bool(creds_file) and Path(str(creds_file)).exists()
    age = int(time.time() - _master_cache["timestamp"]) if _master_cache["timestamp"] else None
    return {
        "configured": configured,
        "file_id": file_id or None,
        "credentials_exists": bool(creds_file) and Path(str(creds_file)).exists(),
        "cache_age_seconds": age,
        "last_error": _master_cache.get("error"),
        "setup_hint": (
            None
            if configured
            else "Chay: python -m auto_fill drive-setup  (xem huong dan ben duoi)"
        ),
    }


@app.get("/master/sheets")
def master_sheets(refresh: bool = False) -> dict[str, Any]:
    """Tra ve danh sach sheet trong file Excel tong tu Google Drive.

    Neu cache con han (< GDRIVE_CACHE_MINUTES phut), tra ve ngay.
    Neu refresh=true, tai lai tu Drive truoc.
    """
    from datetime import datetime

    import openpyxl

    xlsx_path, timestamp, warning = _get_master_xlsx(force=refresh)
    if not xlsx_path:
        raise HTTPException(
            status_code=503, detail=warning or "Khong the tai file tu Google Drive."
        )

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()

    return {
        "sheets": sheets,
        "cached_at": datetime.fromtimestamp(timestamp).isoformat() if timestamp else None,
        "warning": warning,
    }


@app.get("/master/sheet/{sheet_name}")
def master_sheet_data(
    sheet_name: str,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
) -> dict[str, Any]:
    """Tra ve du lieu phan trang cua 1 sheet trong file Excel tong."""
    import openpyxl

    xlsx_path, _, warning = _get_master_xlsx()
    if not xlsx_path:
        raise HTTPException(status_code=503, detail=warning or "Khong the tai file.")

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' khong ton tai.")

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return {
            "columns": [],
            "rows": [],
            "total_rows": 0,
            "page": page,
            "page_size": page_size,
            "total_pages": 0,
            "warning": warning,
        }

    headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(all_rows[0])]
    data_rows = all_rows[1:]
    total_rows = len(data_rows)
    total_pages = max(1, (total_rows + page_size - 1) // page_size)

    start = (page - 1) * page_size
    page_data = data_rows[start : start + page_size]

    rows = []
    for row in page_data:
        row_dict: dict[str, Any] = {}
        for i, val in enumerate(row):
            col = headers[i] if i < len(headers) else f"Col{i}"
            row_dict[col] = val.isoformat() if hasattr(val, "isoformat") else val
        rows.append(row_dict)

    return {
        "columns": headers,
        "rows": rows,
        "total_rows": total_rows,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "warning": warning,
    }


@app.post("/master/refresh")
def master_refresh() -> dict[str, Any]:
    """Buoc tai lai file tu Google Drive (bo qua cache)."""
    from datetime import datetime

    xlsx_path, timestamp, error = _get_master_xlsx(force=True)
    if error and not xlsx_path:
        raise HTTPException(status_code=503, detail=error)
    return {
        "refreshed": True,
        "cached_at": datetime.fromtimestamp(timestamp).isoformat() if timestamp else None,
        "warning": error,
    }


# --------------------------------------------------------------------------- #
# Helpers                                                                      #
# --------------------------------------------------------------------------- #


def _export_response(sheet_aliases: list[str], filename: str) -> StreamingResponse:
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    engine = _get_engine()
    with Session(engine) as s:
        for alias in sheet_aliases:
            model_cls = _ALIAS_TO_MODEL[alias]
            rows = s.query(model_cls).all()
            mapper = model_cls.__mapper__
            col_keys = [col.key for col in mapper.columns]

            ws = wb.create_sheet(title=alias)
            ws.append(col_keys)
            for row in rows:
                ws.append([_serialize(getattr(row, k, None)) for k in col_keys])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _serialize(value: Any) -> Any:
    if isinstance(value, date):
        return value.isoformat()
    return value


def _mask_url(url: str) -> str:
    if "@" in url:
        scheme_user, rest = url.split("@", 1)
        if ":" in scheme_user:
            parts = scheme_user.rsplit(":", 1)
            return f"{parts[0]}:***@{rest}"
    return url
