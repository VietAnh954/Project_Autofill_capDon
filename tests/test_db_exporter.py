"""Unit tests cho filler/db_exporter.py."""

from __future__ import annotations

from collections.abc import Generator
from datetime import date
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy.orm import Session

from auto_fill.db.engine import create_tables, get_engine
from auto_fill.db.models import TravelInsurance
from auto_fill.filler.db_exporter import export_daily_snapshot


@pytest.fixture()
def db_with_data(tmp_path: Path) -> Generator[str, None, None]:
    get_engine.cache_clear()
    db_path = str(tmp_path / "test.db")
    engine = get_engine(db_path)
    create_tables(engine)

    rows = [
        TravelInsurance(
            issued_date=date(2026, 5, 1),
            insured_name="Nguyen Van A",
            insured_id_number="111111111",
            trip_start=date(2026, 6, 1),
            trip_end=date(2026, 6, 10),
            destination="Japan",
        ),
        TravelInsurance(
            issued_date=date(2026, 5, 2),
            insured_name="Tran Thi B",
            insured_id_number="222222222",
            trip_start=date(2026, 7, 1),
            trip_end=date(2026, 7, 5),
            destination="Korea",
        ),
    ]
    with Session(engine) as s:
        s.add_all(rows)
        s.commit()

    yield db_path
    get_engine.cache_clear()


class TestExportDailySnapshot:
    def test_creates_excel_file(self, tmp_path: Path, db_with_data: str) -> None:
        out_path = export_daily_snapshot(
            db_path=db_with_data,
            output_dir=tmp_path / "exports",
            export_date=date(2026, 5, 14),
        )
        assert out_path.exists()
        assert out_path.name == "export_2026-05-14.xlsx"

    def test_excel_has_6_sheets(self, tmp_path: Path, db_with_data: str) -> None:
        out_path = export_daily_snapshot(
            db_path=db_with_data,
            output_dir=tmp_path / "exports",
        )
        wb = openpyxl.load_workbook(out_path)
        assert len(wb.sheetnames) == 6

    def test_travel_sheet_has_data(self, tmp_path: Path, db_with_data: str) -> None:
        out_path = export_daily_snapshot(
            db_path=db_with_data,
            output_dir=tmp_path / "exports",
        )
        wb = openpyxl.load_workbook(out_path)
        ws = wb["Du lich"]
        assert ws.max_row == 3  # header + 2 data rows

    def test_export_dir_created_if_missing(self, tmp_path: Path, db_with_data: str) -> None:
        new_dir = tmp_path / "nested" / "exports"
        assert not new_dir.exists()
        export_daily_snapshot(db_path=db_with_data, output_dir=new_dir)
        assert new_dir.exists()

    def test_uses_today_if_no_date(self, tmp_path: Path, db_with_data: str) -> None:
        from datetime import date as date_cls

        out_path = export_daily_snapshot(
            db_path=db_with_data,
            output_dir=tmp_path / "exports",
        )
        today_str = date_cls.today().isoformat()
        assert today_str in out_path.name
