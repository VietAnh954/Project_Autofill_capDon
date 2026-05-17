"""Unit tests cho Excel reader — dùng fixture tạo bằng openpyxl."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from auto_fill.reader.excel_reader import (
    _build_rename_map,
    _detect_header_row,
    _filldown_buyer,
    read_excel,
)
from auto_fill.utils.errors import ReaderError

ALIASES: dict[str, list[str]] = {
    "insured_name": ["Họ và tên", "ho va ten", "tên khách hàng", "tên ndbh"],
    "insured_id_number": ["CCCD/CMND", "cccd", "số cccd", "cmnd"],
    "insured_dob": ["Ngày sinh", "ngay sinh", "dob"],
    "insured_phone": ["Số điện thoại", "sdt", "phone"],
    "premium": ["Phí bảo hiểm", "phi bao hiem", "tổng phí"],
    "trip_start": ["Ngày đi", "ngay di", "from date"],
    "trip_end": ["Ngày về", "ngay ve", "to date"],
    "destination": ["Nơi đến", "noi den", "destination"],
}


@pytest.fixture
def sample_dulich_xlsx(tmp_path: Path) -> Path:
    """Tạo file Excel mẫu có header tiếng Việt (dạng đại lý gửi)."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Row 1: header
    headers = [
        "Họ và tên",
        "CCCD/CMND",
        "Ngày sinh",
        "Số điện thoại",
        "Ngày đi",
        "Ngày về",
        "Nơi đến",
        "Phí bảo hiểm",
    ]
    ws.append(headers)

    # Row 2-4: data
    ws.append(
        [
            "NGUYỄN VĂN A",
            "012345678901",
            "01/01/1990",
            "0901234567",
            "20/05/2026",
            "27/05/2026",
            "Thái Lan",
            "1500000",
        ]
    )
    ws.append(
        [
            "TRẦN THỊ B",
            "098765432100",
            "15/06/1985",
            "0912345678",
            "01/06/2026",
            "10/06/2026",
            "Nhật Bản",
            "3000000",
        ]
    )
    ws.append(
        [
            "LÊ VĂN C",
            "001122334455",
            "30/12/2000",
            "0933333333",
            "10/07/2026",
            "20/07/2026",
            "Hàn Quốc",
            "2000000",
        ]
    )

    dest = tmp_path / "sample_dulich.xlsx"
    wb.save(dest)
    return dest


@pytest.fixture
def sample_with_empty_rows_xlsx(tmp_path: Path) -> Path:
    """File Excel có 2 hàng trống trước header (thực tế gặp nhiều)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, None, None])  # empty row
    ws.append([None, None, None])  # empty row
    ws.append(["Họ và tên", "CCCD/CMND", "Phí bảo hiểm"])
    ws.append(["NGUYỄN VĂN A", "012345678901", "1500000"])

    dest = tmp_path / "with_empty_rows.xlsx"
    wb.save(dest)
    return dest


class TestReadExcel:
    def test_returns_dataframe_with_correct_rows(self, sample_dulich_xlsx: Path) -> None:
        df = read_excel(sample_dulich_xlsx, ALIASES)
        assert isinstance(df, pd.DataFrame)
        assert len(df) == 3

    def test_columns_renamed_to_canonical(self, sample_dulich_xlsx: Path) -> None:
        df = read_excel(sample_dulich_xlsx, ALIASES)
        assert "insured_name" in df.columns
        assert "insured_id_number" in df.columns
        assert "insured_dob" in df.columns
        assert "premium" in df.columns

    def test_data_values_preserved(self, sample_dulich_xlsx: Path) -> None:
        df = read_excel(sample_dulich_xlsx, ALIASES)
        assert df.iloc[0]["insured_name"] == "NGUYỄN VĂN A"
        assert df.iloc[1]["insured_id_number"] == "098765432100"

    def test_handles_empty_rows_before_header(self, sample_with_empty_rows_xlsx: Path) -> None:
        df = read_excel(sample_with_empty_rows_xlsx, ALIASES)
        assert len(df) == 1
        assert "insured_name" in df.columns

    def test_raises_reader_error_on_corrupt_file(self, tmp_path: Path) -> None:
        bad_file = tmp_path / "corrupt.xlsx"
        bad_file.write_bytes(b"not an excel file")
        with pytest.raises(ReaderError):
            read_excel(bad_file, ALIASES)

    def test_raises_file_not_found(self, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            read_excel(tmp_path / "missing.xlsx", ALIASES)

    def test_empty_aliases_keeps_original_columns(self, sample_dulich_xlsx: Path) -> None:
        df = read_excel(sample_dulich_xlsx, {})
        assert "Họ và tên" in df.columns
        assert "CCCD/CMND" in df.columns


class TestDetectHeaderRow:
    def test_finds_row_with_most_text(self) -> None:
        import pandas as pd

        df = pd.DataFrame(
            [
                [None, None, None],
                ["Họ và tên", "CCCD", "Phí"],
                ["NGUYỄN A", "123456789012", "1500000"],
            ]
        )
        assert _detect_header_row(df) == 1

    def test_first_row_is_header_by_default(self) -> None:
        import pandas as pd

        df = pd.DataFrame(
            [
                ["Họ và tên", "CCCD", "Phí"],
                ["NGUYỄN A", "123456789012", "1500000"],
            ]
        )
        assert _detect_header_row(df) == 0


@pytest.fixture
def buyer_filldown_xlsx(tmp_path: Path) -> Path:
    """Excel with 1 BMBH row + 2 NĐBH rows (buyer cols empty on row 2)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["buyer_name", "buyer_id_number", "insured_name"])
    ws.append(["Nguyễn Thị A", "123456789012", "Con 1"])
    ws.append([None, None, "Con 2"])
    dest = tmp_path / "filldown.xlsx"
    wb.save(dest)
    return dest


class TestFilldownBuyer:
    def test_filldown_propagates_buyer_name(self, buyer_filldown_xlsx: Path) -> None:
        df = read_excel(buyer_filldown_xlsx, {})
        assert df.iloc[1]["buyer_name"] == "Nguyễn Thị A"

    def test_filldown_propagates_buyer_id(self, buyer_filldown_xlsx: Path) -> None:
        df = read_excel(buyer_filldown_xlsx, {})
        assert df.iloc[1]["buyer_id_number"] == "123456789012"

    def test_filldown_does_not_overwrite_existing_value(self) -> None:
        df = pd.DataFrame(
            {
                "buyer_name": ["Alice", "Bob"],
                "insured_name": ["X", "Y"],
            }
        )
        result = _filldown_buyer(df)
        assert result.iloc[1]["buyer_name"] == "Bob"

    def test_filldown_skips_missing_buyer_cols(self) -> None:
        df = pd.DataFrame({"insured_name": ["X", "Y"]})
        result = _filldown_buyer(df)
        assert list(result.columns) == ["insured_name"]

    def test_filldown_replaces_empty_string(self) -> None:
        df = pd.DataFrame({"buyer_name": ["Alice", ""], "insured_name": ["X", "Y"]})
        result = _filldown_buyer(df)
        assert result.iloc[1]["buyer_name"] == "Alice"

    def test_source_buyer_group_id_added(self) -> None:
        df = pd.DataFrame({"buyer_name": ["Alice", None, None], "insured_name": ["X", "Y", "Z"]})
        result = _filldown_buyer(df)
        assert "source_buyer_group_id" in result.columns

    def test_pattern3_group_shares_same_uuid(self) -> None:
        df = pd.DataFrame({"buyer_name": ["Alice", None, None], "insured_name": ["X", "Y", "Z"]})
        result = _filldown_buyer(df)
        ids = result["source_buyer_group_id"].tolist()
        assert ids[0] == ids[1] == ids[2]

    def test_independent_rows_get_different_uuids(self) -> None:
        df = pd.DataFrame({"buyer_name": ["Alice", "Bob"], "insured_name": ["X", "Y"]})
        result = _filldown_buyer(df)
        ids = result["source_buyer_group_id"].tolist()
        assert ids[0] != ids[1]

    def test_buyer_name_filled_after_uuid_tagging(self) -> None:
        df = pd.DataFrame({"buyer_name": ["Alice", None, None], "insured_name": ["X", "Y", "Z"]})
        result = _filldown_buyer(df)
        assert result["buyer_name"].tolist() == ["Alice", "Alice", "Alice"]


class TestPattern3Acceptance:
    """8.4 acceptance: sample_du_lich.xlsx gia dinh Pham Minh Cuong share same group id."""

    SAMPLE_PATH = Path("sample_capdon/sample_du_lich.xlsx")

    def test_pham_minh_cuong_rows_share_group_id(self) -> None:
        from auto_fill.mapper.aliases_loader import load_aliases

        aliases = load_aliases()
        df = read_excel(self.SAMPLE_PATH, aliases)
        group = df[df["buyer_name"] == "Phạm Minh Cường"]
        assert len(group) >= 3, "Expected at least 3 rows for Pham Minh Cuong group"
        group_ids = group["source_buyer_group_id"].unique()
        assert len(group_ids) == 1, f"Expected 1 unique group ID, got {len(group_ids)}"

    def test_different_buyers_have_different_group_ids(self) -> None:
        from auto_fill.mapper.aliases_loader import load_aliases

        aliases = load_aliases()
        df = read_excel(self.SAMPLE_PATH, aliases)
        cuong_id = df[df["buyer_name"] == "Phạm Minh Cường"]["source_buyer_group_id"].iloc[0]
        an_id = df[df["buyer_name"] == "Nguyễn Văn An"]["source_buyer_group_id"].iloc[0]
        assert cuong_id != an_id


class TestBuildRenameMap:
    def test_exact_match_case_insensitive(self) -> None:
        rename_map = _build_rename_map(["Họ và tên", "CCCD/CMND"], ALIASES)
        assert rename_map.get("Họ và tên") == "insured_name"
        assert rename_map.get("CCCD/CMND") == "insured_id_number"

    def test_fuzzy_match_for_typo(self) -> None:
        rename_map = _build_rename_map(["ho va ten"], ALIASES)
        assert rename_map.get("ho va ten") == "insured_name"

    def test_unmatched_column_not_in_rename_map(self) -> None:
        rename_map = _build_rename_map(["Cột không biết"], ALIASES)
        assert "Cột không biết" not in rename_map
