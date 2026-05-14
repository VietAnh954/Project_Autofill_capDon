"""Unit tests cho CLI rollback va replay commands."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from click.testing import CliRunner

from auto_fill.__main__ import replay, rollback


@pytest.fixture
def master_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Ngày", "STT", "Họ Và Tên"])
    ws.append(["2026-05-14", 1, "Nguyen Van A"])
    wb.save(path)
    return path


@pytest.fixture
def backup_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "backup.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Ngày", "STT", "Họ Và Tên"])
    wb.save(path)
    return path


def _make_settings(tmp_path: Path, master_path: Path) -> MagicMock:
    s = MagicMock()
    s.master_file_path = master_path
    s.data_root = tmp_path
    s.dry_run = False
    s.log_level = "INFO"
    s.log_dir = tmp_path / "logs"
    return s


class TestRollbackCommand:
    def test_rollback_with_yes_flag_succeeds(
        self, tmp_path: Path, master_xlsx: Path, backup_xlsx: Path
    ) -> None:
        settings = _make_settings(tmp_path, master_xlsx)
        runner = CliRunner()

        with patch("auto_fill.filler.backup.snapshot"):
            result = runner.invoke(
                rollback,
                ["--to", str(backup_xlsx), "--yes"],
                obj={"settings": settings},
            )

        assert result.exit_code == 0, result.output
        assert "[rollback]" in result.output

    def test_rollback_without_yes_aborts_on_n(
        self, tmp_path: Path, master_xlsx: Path, backup_xlsx: Path
    ) -> None:
        settings = _make_settings(tmp_path, master_xlsx)
        runner = CliRunner()

        result = runner.invoke(
            rollback,
            ["--to", str(backup_xlsx)],
            input="n\n",
            obj={"settings": settings},
        )

        assert result.exit_code != 0 or "Aborted" in result.output

    def test_rollback_copies_backup_over_master(
        self, tmp_path: Path, master_xlsx: Path, backup_xlsx: Path
    ) -> None:
        settings = _make_settings(tmp_path, master_xlsx)
        runner = CliRunner()

        with patch("auto_fill.filler.backup.snapshot"):
            runner.invoke(
                rollback,
                ["--to", str(backup_xlsx), "--yes"],
                obj={"settings": settings},
            )

        wb = openpyxl.load_workbook(master_xlsx)
        ws = wb.active
        assert ws.max_row == 1  # backup had only header row, master had header+1 row


class TestReplayCommand:
    def test_replay_empty_folder(self, tmp_path: Path, master_xlsx: Path) -> None:
        inbox = tmp_path / "inbox"
        inbox.mkdir()
        settings = _make_settings(tmp_path, master_xlsx)
        runner = CliRunner()

        result = runner.invoke(
            replay,
            ["--folder", str(inbox)],
            obj={"settings": settings},
        )

        assert result.exit_code == 0
        assert "Khong co file nao" in result.output

    def test_replay_default_folder_is_inbox(self, tmp_path: Path, master_xlsx: Path) -> None:
        inbox = tmp_path / "inbox"
        inbox.mkdir()
        settings = _make_settings(tmp_path, master_xlsx)
        runner = CliRunner()

        result = runner.invoke(
            replay,
            [],
            obj={"settings": settings},
        )

        assert result.exit_code == 0
        assert "Khong co file nao" in result.output

    def test_replay_unsupported_extension_ignored(self, tmp_path: Path, master_xlsx: Path) -> None:
        folder = tmp_path / "replay_test"
        folder.mkdir()
        (folder / "readme.txt").write_text("ignore me")
        settings = _make_settings(tmp_path, master_xlsx)
        runner = CliRunner()

        result = runner.invoke(
            replay,
            ["--folder", str(folder)],
            obj={"settings": settings},
        )

        assert result.exit_code == 0
        assert "Khong co file nao" in result.output
