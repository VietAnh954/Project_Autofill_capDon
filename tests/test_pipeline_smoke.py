"""End-to-end smoke test cho pipeline Du lich.

Test nay KHONG can Outlook chay that. Toan bo COM duoc mock.
Sau khi chay, file master co them dung 1 dong cuoi sheet Du lich.
"""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock

import openpyxl

from auto_fill.filler.backup import snapshot
from auto_fill.filler.excel_filler import append_travel_rows
from auto_fill.mail.marker import mark_processed
from auto_fill.mapper.classifier import classify
from auto_fill.mapper.dedup import is_duplicate
from auto_fill.mapper.validator import validate_and_raise

_SHEET = "Du lịch"
_HEADERS = [
    "issued_date",
    "STT",
    "insured_name",
    "insured_dob",
    "insured_id_number",
    "effective_from",
    "trip_start",
    "trip_end",
    "trip_days",
    "destination",
]

_RECORD = {
    "issued_date": date(2026, 5, 14),
    "insured_name": "NGUYEN VAN A",
    "insured_dob": date(1990, 1, 1),
    "insured_id_number": "012345678901",
    "trip_start": date(2026, 5, 20),
    "trip_end": date(2026, 5, 27),
    "destination": "Thai Lan",
}


def _make_master(tmp_path: Path) -> Path:
    path = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _SHEET
    ws.append(_HEADERS)
    wb.save(path)
    return path


def _make_csv(tmp_path: Path) -> Path:
    path = tmp_path / "phieu_dulich.csv"
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "issued_date",
                "insured_name",
                "insured_dob",
                "insured_id_number",
                "trip_start",
                "trip_end",
                "destination",
            ],
        )
        writer.writeheader()
        writer.writerow(
            {
                "issued_date": "14/05/2026",
                "insured_name": "NGUYEN VAN A",
                "insured_dob": "01/01/1990",
                "insured_id_number": "012345678901",
                "trip_start": "20/05/2026",
                "trip_end": "27/05/2026",
                "destination": "Thai Lan",
            }
        )
    return path


class TestPipelineSmoke:
    def test_backup_created(self, tmp_path: Path) -> None:
        master = _make_master(tmp_path)
        backup_root = tmp_path / "backup"
        dest = snapshot(master, backup_root, today=date(2026, 5, 14))
        assert dest.exists()

    def test_classifier_returns_travel(self) -> None:
        sheet = classify("Phieu cap don du lich", "agent@example.com")
        assert sheet.alias == "travel"

    def test_validate_and_raise_passes(self) -> None:
        validate_and_raise(_RECORD, "travel")

    def test_dedup_false_on_empty_master(self, tmp_path: Path) -> None:
        master = _make_master(tmp_path)
        assert is_duplicate(_RECORD, master, sheet_name=_SHEET) is False

    def test_append_row_success(self, tmp_path: Path) -> None:
        master = _make_master(tmp_path)
        written = append_travel_rows([_RECORD], master, sheet_name=_SHEET)
        assert written == 1
        wb = openpyxl.load_workbook(master)
        ws = wb[_SHEET]
        assert ws.max_row == 2
        assert ws.cell(row=2, column=2).value == 1  # STT

    def test_full_pipeline_csv_to_master(self, tmp_path: Path) -> None:
        """Smoke test: 1 CSV file -> 1 dong moi trong master."""
        import contextlib

        from auto_fill.mapper.normalizer import normalize_date
        from auto_fill.reader.csv_reader import read_csv

        master = _make_master(tmp_path)
        csv_path = _make_csv(tmp_path)

        df = read_csv(csv_path, aliases={})
        assert len(df) == 1

        row = df.iloc[0].to_dict()
        for f in ("issued_date", "insured_dob", "trip_start", "trip_end"):
            if row.get(f) is not None:
                with contextlib.suppress(Exception):
                    row[f] = normalize_date(row[f])

        validate_and_raise(row, "travel")
        assert not is_duplicate(row, master, sheet_name=_SHEET)
        written = append_travel_rows([row], master, sheet_name=_SHEET)
        assert written == 1

        wb = openpyxl.load_workbook(master)
        ws = wb[_SHEET]
        assert ws.max_row == 2
        assert ws.cell(row=2, column=3).value == "NGUYEN VAN A"

    def test_dedup_true_after_insert(self, tmp_path: Path) -> None:
        master = _make_master(tmp_path)
        append_travel_rows([_RECORD], master, sheet_name=_SHEET)
        assert is_duplicate(_RECORD, master, sheet_name=_SHEET, header_row=0) is True

    def test_mark_processed_mock(self) -> None:
        ns = MagicMock()
        inbox = MagicMock()
        inbox.Folders.Count = 0
        ns.GetDefaultFolder.return_value = inbox
        mail = MagicMock()
        mail.UnRead = True
        ns.GetItemFromID.return_value = mail
        result = mark_processed(ns, "AAA", "Processed")
        assert result is True
        assert mail.UnRead is False

    def test_multiple_records_appended(self, tmp_path: Path) -> None:
        master = _make_master(tmp_path)
        record2 = {**_RECORD, "insured_id_number": "099999999999"}
        written = append_travel_rows([_RECORD, record2], master, sheet_name=_SHEET)
        assert written == 2
        wb = openpyxl.load_workbook(master)
        ws = wb[_SHEET]
        assert ws.max_row == 3
