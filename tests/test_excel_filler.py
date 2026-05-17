"""Unit tests cho Excel filler."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import ClassVar

import openpyxl
import pytest

from auto_fill.filler.excel_filler import _next_stt, append_rows, append_travel_rows
from auto_fill.mapper.sheet_mapping import STUDENT

_SHEET = "Du lịch"
_HEADERS = [
    "Ngày",
    "STT",
    "Họ Và Tên",
    "Ngày sinh",
    "CCCD/CMND/Paspost ID",
    "Ngày thanh toán",
    "Ngày đi",
    "Ngày về",
    "Số ngày",
    "Nơi đến",
    "Phạm vi",
    "Gói tham gia",
    "Plan tham gia",
    "Phí bảo hiểm",
    "Họ tên người mua",
]

_RECORD = {
    "issued_date": date(2026, 5, 14),
    "insured_name": "NGUYEN VAN A",
    "insured_dob": date(1990, 1, 1),
    "insured_id_number": "012345678901",
    "trip_start": date(2026, 5, 20),
    "trip_end": date(2026, 5, 27),
    "destination": "Thai Lan",
    "scope": "Chau A",
    "plan": "Plan A",
    "premium": 1500000.0,
    "buyer_name": "NGUYEN VAN B",
}


def _make_workbook(tmp_path: Path, existing_rows: int = 0) -> Path:
    path = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _SHEET
    ws.append(_HEADERS)
    for i in range(1, existing_rows + 1):
        ws.append(
            [
                date(2026, 5, 1),
                i,
                f"Person {i}",
                date(1990, 1, 1),
                f"00000000000{i}",
                None,
                date(2026, 5, 10),
                date(2026, 5, 17),
                7,
                "Singapore",
                "Chau A",
                None,
                "Plan B",
                2000000.0,
                f"Buyer {i}",
            ]
        )
    wb.save(path)
    return path


class TestAppendTravelRows:
    def test_appends_one_row(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path)
        written = append_travel_rows([_RECORD], path)
        assert written == 1

    def test_row_count_increases(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path)
        wb_before = openpyxl.load_workbook(path)
        rows_before = wb_before[_SHEET].max_row
        append_travel_rows([_RECORD], path)
        wb_after = openpyxl.load_workbook(path)
        assert wb_after[_SHEET].max_row == rows_before + 1

    def test_appends_multiple_rows(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path)
        written = append_travel_rows([_RECORD, _RECORD], path)
        assert written == 2

    def test_insured_name_written(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path)
        append_travel_rows([_RECORD], path)
        wb = openpyxl.load_workbook(path)
        ws = wb[_SHEET]
        row = ws.max_row
        assert ws.cell(row=row, column=3).value == "NGUYEN VAN A"

    def test_stt_auto_increments(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path, existing_rows=5)
        append_travel_rows([_RECORD], path)
        wb = openpyxl.load_workbook(path)
        ws = wb[_SHEET]
        row = ws.max_row
        assert ws.cell(row=row, column=2).value == 6

    def test_stt_starts_at_1_for_empty_sheet(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path, existing_rows=0)
        append_travel_rows([_RECORD], path)
        wb = openpyxl.load_workbook(path)
        ws = wb[_SHEET]
        assert ws.cell(row=2, column=2).value == 1

    def test_trip_days_computed(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path)
        append_travel_rows([_RECORD], path)
        wb = openpyxl.load_workbook(path)
        ws = wb[_SHEET]
        row = ws.max_row
        assert ws.cell(row=row, column=9).value == 7  # 27-20 = 7

    def test_issued_date_written(self, tmp_path: Path) -> None:
        from datetime import datetime

        path = _make_workbook(tmp_path)
        append_travel_rows([_RECORD], path)
        wb = openpyxl.load_workbook(path)
        ws = wb[_SHEET]
        row = ws.max_row
        val = ws.cell(row=row, column=1).value
        # openpyxl reads date cells back as datetime
        if isinstance(val, datetime):
            val = val.date()
        assert val == date(2026, 5, 14)

    def test_file_not_found_raises(self, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            append_travel_rows([_RECORD], tmp_path / "missing.xlsx")

    def test_wrong_sheet_name_raises(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path)
        with pytest.raises(KeyError):
            append_travel_rows([_RECORD], path, sheet_name="Wrong Sheet")

    def test_optional_field_none_skipped(self, tmp_path: Path) -> None:
        record = {**_RECORD}
        del record["buyer_name"]
        path = _make_workbook(tmp_path)
        written = append_travel_rows([record], path)
        assert written == 1

    def test_premium_written_as_float(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path)
        append_travel_rows([_RECORD], path)
        wb = openpyxl.load_workbook(path)
        ws = wb[_SHEET]
        row = ws.max_row
        assert ws.cell(row=row, column=14).value == 1500000.0


class TestPattern3BuyerEmpty:
    """8.6: buyer cols empty on continuation rows sharing source_buyer_group_id."""

    _HSSV_HEADERS: ClassVar[list[str]] = [
        "Ngày",  # col 1 issued_date
        "Họ tên NĐBH",  # col 2 insured_name
        "Ngày sinh NĐBH",  # col 3 insured_dob
        "Giới tính",  # col 4
        "CCCD/CMND",  # col 5
        "Email",  # col 6
        "SĐT",  # col 7
        "Địa chỉ",  # col 8
        "Trường",  # col 9
        "Lớp",  # col 10
        "Họ tên BMBH",  # col 11 buyer_name
        "Quan hệ",  # col 12 buyer_relation
        "Ngày sinh BMBH",  # col 13 buyer_dob
        "CCCD BMBH",  # col 14 buyer_id_number
        "Email BMBH",  # col 15 buyer_email
        "SĐT BMBH",  # col 16 buyer_phone
        "Địa chỉ BMBH",  # col 17 buyer_address
    ]

    _GROUP_ID = "test-group-uuid-1234"

    _BASE_BUYER: ClassVar[dict[str, str]] = {
        "buyer_name": "Lê Thị Hoa",
        "buyer_relation": "Mẹ",
        "buyer_dob": "1980-01-01",
        "buyer_id_number": "011111111111",
        "buyer_email": "hoa@example.com",
        "buyer_phone": "0901111111",
        "buyer_address": "HCM",
    }

    def _make_hssv_workbook(self, tmp_path: Path) -> Path:
        path = tmp_path / "master_hssv.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HSSV"
        ws.append(self._HSSV_HEADERS)
        wb.save(path)
        return path

    def test_first_row_has_buyer_name(self, tmp_path: Path) -> None:
        path = self._make_hssv_workbook(tmp_path)
        records = [
            {"insured_name": "Con 1", "source_buyer_group_id": self._GROUP_ID, **self._BASE_BUYER},
            {"insured_name": "Con 2", "source_buyer_group_id": self._GROUP_ID},
        ]
        append_rows(records, STUDENT, path, "HSSV")
        wb = openpyxl.load_workbook(path)
        ws = wb["HSSV"]
        assert ws.cell(row=2, column=11).value == "Lê Thị Hoa"

    def test_continuation_row_has_empty_buyer_name(self, tmp_path: Path) -> None:
        path = self._make_hssv_workbook(tmp_path)
        records = [
            {"insured_name": "Con 1", "source_buyer_group_id": self._GROUP_ID, **self._BASE_BUYER},
            {"insured_name": "Con 2", "source_buyer_group_id": self._GROUP_ID},
        ]
        append_rows(records, STUDENT, path, "HSSV")
        wb = openpyxl.load_workbook(path)
        ws = wb["HSSV"]
        assert ws.cell(row=3, column=11).value is None

    def test_all_buyer_cols_empty_for_continuation(self, tmp_path: Path) -> None:
        path = self._make_hssv_workbook(tmp_path)
        records = [
            {"insured_name": "Con 1", "source_buyer_group_id": self._GROUP_ID, **self._BASE_BUYER},
            {"insured_name": "Con 2", "source_buyer_group_id": self._GROUP_ID},
            {"insured_name": "Con 3", "source_buyer_group_id": self._GROUP_ID},
        ]
        append_rows(records, STUDENT, path, "HSSV")
        wb = openpyxl.load_workbook(path)
        ws = wb["HSSV"]
        # rows 3 and 4 are continuation rows; buyer cols 11-17 must all be None
        for row_idx in (3, 4):
            for col_idx in range(11, 18):
                val = ws.cell(row=row_idx, column=col_idx).value
                assert val is None, f"Row {row_idx} col {col_idx}: expected None, got {val!r}"

    def test_independent_rows_each_write_buyer(self, tmp_path: Path) -> None:
        path = self._make_hssv_workbook(tmp_path)
        records = [
            {"insured_name": "A", "source_buyer_group_id": "group-1", **self._BASE_BUYER},
            {"insured_name": "B", "source_buyer_group_id": "group-2", "buyer_name": "Other Buyer"},
        ]
        append_rows(records, STUDENT, path, "HSSV")
        wb = openpyxl.load_workbook(path)
        ws = wb["HSSV"]
        assert ws.cell(row=2, column=11).value == "Lê Thị Hoa"
        assert ws.cell(row=3, column=11).value == "Other Buyer"

    def test_no_group_id_always_writes_buyer(self, tmp_path: Path) -> None:
        path = self._make_hssv_workbook(tmp_path)
        records = [
            {"insured_name": "A", **self._BASE_BUYER},
            {"insured_name": "B", **self._BASE_BUYER},
        ]
        append_rows(records, STUDENT, path, "HSSV")
        wb = openpyxl.load_workbook(path)
        ws = wb["HSSV"]
        assert ws.cell(row=2, column=11).value == "Lê Thị Hoa"
        assert ws.cell(row=3, column=11).value == "Lê Thị Hoa"

    def test_travel_pattern3_skips_buyer_name_on_continuation(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path)
        records = [
            {**_RECORD, "source_buyer_group_id": "grp", "buyer_name": "Parent"},
            {**_RECORD, "source_buyer_group_id": "grp", "insured_name": "Child"},
        ]
        append_travel_rows(records, path)
        wb = openpyxl.load_workbook(path)
        ws = wb[_SHEET]
        # row 2 = first record (buyer_name col 15 = "Parent")
        assert ws.cell(row=2, column=15).value == "Parent"
        # row 3 = continuation (buyer_name col 15 = None)
        assert ws.cell(row=3, column=15).value is None


class TestNextStt:
    def test_empty_sheet_returns_1(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path, existing_rows=0)
        wb = openpyxl.load_workbook(path)
        ws = wb[_SHEET]
        assert _next_stt(ws, col_idx=2, header_row=1) == 1

    def test_returns_max_plus_1(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path, existing_rows=3)
        wb = openpyxl.load_workbook(path)
        ws = wb[_SHEET]
        assert _next_stt(ws, col_idx=2, header_row=1) == 4
