"""Unit tests cho CLI export command."""

from __future__ import annotations

from collections.abc import Generator
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest
from click.testing import CliRunner
from sqlalchemy.orm import Session

from auto_fill.__main__ import export_cmd
from auto_fill.db.engine import create_tables, get_engine
from auto_fill.db.models import TravelInsurance


@pytest.fixture()
def db_with_travel(tmp_path: Path) -> Generator[str, None, None]:
    get_engine.cache_clear()
    db_path = str(tmp_path / "test.db")
    engine = get_engine(db_path)
    create_tables(engine)

    with Session(engine) as s:
        s.add(
            TravelInsurance(
                issued_date=date(2026, 5, 1),
                insured_name="Nguyen Van A",
                insured_id_number="111111111",
                trip_start=date(2026, 6, 1),
                trip_end=date(2026, 6, 10),
                destination="Japan",
            )
        )
        s.commit()

    yield db_path
    get_engine.cache_clear()


def _make_settings(tmp_path: Path, db_path: str) -> MagicMock:
    s = MagicMock()
    s.db_path = Path(db_path)
    s.data_root = tmp_path
    s.dry_run = False
    s.log_level = "INFO"
    s.log_dir = tmp_path / "logs"
    return s


class TestExportCommand:
    def test_export_all_creates_xlsx(self, tmp_path: Path, db_with_travel: str) -> None:
        settings = _make_settings(tmp_path, db_with_travel)
        runner = CliRunner()
        result = runner.invoke(
            export_cmd,
            [],
            obj={"settings": settings},
        )
        assert result.exit_code == 0, result.output
        assert "[export]" in result.output

    def test_export_single_sheet(self, tmp_path: Path, db_with_travel: str) -> None:
        out_file = str(tmp_path / "travel.xlsx")
        settings = _make_settings(tmp_path, db_with_travel)
        runner = CliRunner()
        result = runner.invoke(
            export_cmd,
            ["--sheet", "travel", "--to", out_file],
            obj={"settings": settings},
        )
        assert result.exit_code == 0, result.output
        assert Path(out_file).exists()

    def test_export_single_sheet_has_data(self, tmp_path: Path, db_with_travel: str) -> None:
        out_file = str(tmp_path / "travel.xlsx")
        settings = _make_settings(tmp_path, db_with_travel)
        runner = CliRunner()
        runner.invoke(
            export_cmd,
            ["--sheet", "travel", "--to", out_file],
            obj={"settings": settings},
        )
        wb = openpyxl.load_workbook(out_file)
        ws = wb.active
        assert ws is not None
        assert ws.max_row >= 2  # header + at least 1 data row

    def test_export_all_has_6_sheets(self, tmp_path: Path, db_with_travel: str) -> None:
        settings = _make_settings(tmp_path, db_with_travel)
        runner = CliRunner()
        result = runner.invoke(
            export_cmd,
            [],
            obj={"settings": settings},
        )
        assert result.exit_code == 0, result.output
        exports_dir = tmp_path / "exports"
        xlsx_files = list(exports_dir.glob("*.xlsx"))
        assert len(xlsx_files) == 1
        wb = openpyxl.load_workbook(xlsx_files[0])
        assert len(wb.sheetnames) == 6
